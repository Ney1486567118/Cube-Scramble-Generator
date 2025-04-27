import time
import os
import base64
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from cairosvg import svg2png

# ====== 可配置项 ======
DRIVER_PATH = './chromedriver.exe'
SCREENSHOT_DIR = 'screenshots'      # 截图暂存路径

# ====== 设置打乱数量 ======
def set_scramble_num():
    print("请在网页中手动选择项目（如“三阶速拧”、“二阶速拧”），选择完毕后按 'Enter'。")
    input()
    print("请输入正式打乱数量：")
    num_main = int(input())
    print("请输入备用打乱数量：")
    num_spare = int(input())
    return num_main, num_spare

# ====== 获取打乱信息并截图（base64 SVG） ======
def capture_scramble(idx, is_spare=False):
    # 获取打乱公式
    scramble_elem = driver.find_element(By.ID, "scrambleTxt")
    scramble_text = scramble_elem.text.strip()

    # 获取图像 base64
    img_elem = driver.find_element(By.XPATH, '//*[@id="toolsDiv"]/div[1]/div/img')
    img_data = img_elem.get_attribute("src")

    assert img_data.startswith("data:image/svg+xml;base64,"), "不是SVG base64格式"
    svg_base64 = img_data.replace("data:image/svg+xml;base64,", "")
    svg_bytes = base64.b64decode(svg_base64)

    # 保存为临时 SVG
    svg_temp_path = "temp_img.svg"
    with open(svg_temp_path, "wb") as f:
        f.write(svg_bytes)

    # 转换为 PNG（适用于 Excel）
    label = f"({idx})" if is_spare else f"{idx}"
    img_filename = os.path.join(SCREENSHOT_DIR, f"{label}.png")
    svg2png(url=svg_temp_path, write_to=img_filename, output_width=300, output_height=225)

    # 写入 Excel
    ws.append([label, scramble_text])

    # 设置字体样式
    row = ws.max_row
    font_style = Font(name="Courier New", size=20)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ["A", "B"]:
        cell = ws[f"{col}{row}"]
        cell.font = font_style
        cell.alignment = alignment_center

    # 设置行高
    ws.row_dimensions[row].height = 120

    # 插入图片
    img = ExcelImage(img_filename)
    img.width = 200
    img.height = 150
    ws.add_image(img, f"C{ws.max_row}")

if __name__ == '__main__':
    # ====== 启动 Selenium ======
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(DRIVER_PATH), options=options)

    # ====== 打开 CSTimer ======
    driver.get("https://cstimer.net/")
    # 等待直到页面加载完成
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )
    # 打开左侧工具栏 → 显示魔方图案
    tools_enable = driver.find_element(By.XPATH, '//*[@id="leftbar"]/div[7]')
    tools_enable.click()


    while True:
    # ====== 设置项目名称 ======
        output_file = input("请输入希望保存的文件名（如：三阶速拧，则打乱将保存为'三阶速拧.xlsx'）（或输入'n'退出程序）：")

        if output_file.lower() == 'n':
            print("打乱生成结束")
            driver.quit()
            break

        # ====== 选择项目 + 设置打乱数量 ======
        num_main, num_spare = set_scramble_num()
        total = num_main + num_spare

        # ====== 创建截图目录 ======
        if not os.path.exists(SCREENSHOT_DIR):
            os.makedirs(SCREENSHOT_DIR)

        # ====== 设置 Excel 表格 ======
        wb = Workbook()
        ws = wb.active
        ws.title = output_file
        # 表头
        ws.append(["序号", "打乱公式", "打乱图案"])
        # 设置表头格式
        header_font = Font(name="Courier New", size=20, bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        for col in ["A", "B", "C"]:
            cell = ws[f"{col}1"]
            cell.font = header_font
            cell.alignment = header_align


        # 设置列宽
        ws.column_dimensions["A"].width = 10     # 编号列
        ws.column_dimensions["B"].width = 70     # 打乱公式列
        ws.column_dimensions["C"].width = 30     # 图案列

        # ====== 生成所有打乱 ======
        for i in range(1, total + 1):
            if i > 1:
                next_btn = driver.find_element(By.XPATH, '//*[@id="scrambleDiv"]//span[contains(text(),"下一条")]')
                next_btn.click()
                time.sleep(2)
            is_spare = (i > num_main)
            capture_scramble(i - num_main if is_spare else i, is_spare)

        # ====== 清理 & 保存 ======
        if os.path.exists("temp_img.svg"):
            os.remove("temp_img.svg")
        wb.save(f"{output_file}.xlsx")
        print(f"\n✅ 打乱生成完成，已保存为 {output_file}.xlsx")
